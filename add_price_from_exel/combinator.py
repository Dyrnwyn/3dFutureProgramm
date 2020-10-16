import re
import os
import shutil
import openpyxl
from PIL import Image, ExifTags

def searchFl(flExt, fldr=""):
    # Во временно директории ищем все jpeg файлы
    # И добавляем их в список
    filteredFlList = []
    with os.scandir(fldr) as flList:
        for fl in flList:
            if not fl.name.startswith('.') and fl.is_file() and \
                    re.findall(r"^.*\." + flExt + "*", fl.name):
                filteredFlList.append(fl.name)
    return filteredFlList

def find_files():
    i = os.scandir('.')
    listofFiles = []
    for element in i:
        if element.is_file():
            listofFiles.append(element.name)
    return listofFiles

def resize_image(fileName):
    baseSize = 1500
    img = Image.open(fileName)
    #info = img._getexif()
    #print (info)
    #for tag, value in info
    x, y = img.size
    if x > y:
        height = baseSize
        width = int(height / y * x)
        return img.resize((width, height), Image.BICUBIC)
    else:
        width = baseSize
        height = int(width / x * y)
        return img.resize((width, height), Image.BICUBIC)

# def renameJPG(xlsxFL, psdFiles):
#     print(xlsxFL)
#     xlFile = openpyxl.load_workbook(xlsxFL, read_only=True)
#     ws = xlFile["Фото"]
#     for f in psdFiles:
#         row = 11
#         while ws["C"+ str(row)].value != None:
#             name = ws["E" + str(row)].value + "_" + ws["C" + str(row)].value + "_" + ws["D" + str(row)].value
#             if name in f:
#                 name_with_price = ""
#                 split_file_name = f.split("_")
#                 split_file_name[8] = ws["K" + str(row)].value
#                 for i in split_file_name:
#                     print(i)
#                     if "psd" in i:
#                         name_with_price = name_with_price + i
#                     else:
#                         name_with_price = name_with_price + i + "_"
#                 os.rename(f, name_with_price)
#                 # if fileName in i:
#                 #     print (fileName)
#                 #     #img = resize_image(i)
#                 #     img = Image.open(i)
#                 #     if i[-4:] == ".png":
#                 #         img.save(fileName + "-" + name + ".png", dpi=(300,300), compress_level=2)
#                 #     else:
#                 #         img.save(fileName + "-" + name + ".jpg", dpi=(300,300), quality=98)
#             row += 1



def renameJPG(txtFL, psdFiles):
    for f in psdFiles:
        file = open(txtFL, 'r')
        for line in file:
            if line[1] == "$":
               split_file_name = line.split(";")
               name = split_file_name[6] +"_"+ split_file_name[4] +"_"+ split_file_name[5]
               if name in f:
                   split_psd_filename = f.split("_")
                   split_psd_filename[8] = split_file_name[10]
                   os.rename(f,"_".join(split_psd_filename)) 
            else:
                pass
        file.close()


        # row = 11
        # while ws["C"+ str(row)].value != None:
        #     name = ws["E" + str(row)].value + "_" + ws["C" + str(row)].value + "_" + ws["D" + str(row)].value
        #     if name in f:
        #         name_with_price = ""
        #         split_file_name = f.split("_")
        #         split_file_name[8] = ws["K" + str(row)].value
        #         for i in split_file_name:
        #             print(i)
        #             if "psd" in i:
        #                 name_with_price = name_with_price + i
        #             else:
        #                 name_with_price = name_with_price + i + "_"
        #         os.rename(f, name_with_price)
        #         # if fileName in i:
        #         #     print (fileName)
        #         #     #img = resize_image(i)
        #         #     img = Image.open(i)
        #         #     if i[-4:] == ".png":
        #         #         img.save(fileName + "-" + name + ".png", dpi=(300,300), compress_level=2)
        #         #     else:
        #         #         img.save(fileName + "-" + name + ".jpg", dpi=(300,300), quality=98)
        #     row += 1

# def listHVPhoto(fl):
#     photoH = []
#     photoV = []
#     for i in fl:
#         img = Image.open(i)
#         if img.size[0] < img.size[1]:
#             photoV.append(i)
#         else:
#             photoH.append(i)
#         img.close()
#     return photoH, photoV

# def listHVPhotoForCrop(fl):
#     photoH = []
#     photoV = []
#     for i in fl:
#         if "H" in i:
#             photoH.append(i)
#         elif "V" in i:
#             photoV.append(i)
#     return photoH, photoV

# def createFileForRemoveBGH(photoH):
#     baseSize = 4000
#     nextPhotoY = 0
#     basImg = Image.new('RGB', (4000, 6250), color = (255, 255, 255))
#     countImg = 0
#     baseFileName = "H_"
#     countLen = 0
#     for i in photoH:
#         print() 
#         countImg += 1
#         img = Image.open(i)
#         x, y = img.size
#         width = baseSize
#         height = int(width / x * y)
#         imgR = img.resize((width, height), Image.BICUBIC)
#         #countLen += 1
#         if countImg == 1:
#             nextPhotoY = height
#             baseFileName += i[:-4] + "_"
#             basImg = Image.new('RGB', (4000, 6250), color = (255, 255, 255))     
#             basImg.paste(imgR, (0, 0))
#             if photoH.index(i) == (len(photoH) - 1):
#                 basImg.save(baseFileName + "NONE_.jpg", dpi=(300,300), quality=98)
#         else:
#             baseFileName += i[:-4] + "_"
#             bottomPoint = nextPhotoY + height
#             basImg.paste(imgR, (0, nextPhotoY))
#             basImg = basImg.crop((0, 0, 4000, bottomPoint))
#             basImg.save(baseFileName + ".jpg", dpi=(300,300), quality=98)
#             baseFileName = "H_"
#             countImg = 0


# def createFileForRemoveBGV(photoV):
#     baseSize = 4000
#     nextPhotoX = 0
#     basImg = Image.new('RGB', (6250, 4000), color = (255, 255, 255))
#     countImg = 0
#     baseFileName = "V_"
#     countLen = 0
#     for i in photoV:
#         countImg += 1
#         img = Image.open(i)
#         x, y = img.size
#         height = baseSize
#         width = int(height / y * x)
#         imgR = img.resize((width, height), Image.BICUBIC)
#         #countLen += 1
#         if countImg == 1:
#             nextPhotoX = width
#             baseFileName += i[:-4] + "_"
#             basImg = Image.new('RGB', (6250, 4000), color = (255, 255, 255))     
#             basImg.paste(imgR, (0, 0))
#             if photoV.index(i) == (len(photoV) - 1):
#                 basImg.save(baseFileName + "NONE_.jpg", dpi=(300,300), quality=98)
#         else:
#             baseFileName += i[:-4] + "_"
#             basImg.paste(imgR, (nextPhotoX, 0))
#             bottomPoint = nextPhotoX + width
#             basImg = basImg.crop((0, 0, bottomPoint, 4000))
#             basImg.save(baseFileName + ".jpg", dpi=(300,300), quality=98)
#             baseFileName = "V_"
#             countImg = 0

# def cropImage(photoH, photoV):
#     for i in photoH:
#         img = Image.open(i)
#         x, y = img.size
#         imgTopName = i.split("_")[1]
#         imgTop = img.crop((0,0,4000,y/2))
#         imgTop.save(imgTopName + ".png")
#         imgBottomName = i.split("_")[2]
#         if "NONE" not in imgBottomName:
#             imgBottom = img.crop((0,y/2,4000,y))
#             imgBottom.save(imgBottomName + ".png", dpi=(300,300), compress_level=2)

#     for i in photoV:
#         img = Image.open(i)
#         x, y = img.size
#         imgTopName = i.split("_")[1]
#         imgTop = img.crop((0,0,x/2,4000))
#         imgTop.save(imgTopName + ".png")
#         imgBottomName = i.split("_")[2]
#         if "NONE" not in imgBottomName:
#             imgBottom = img.crop((x/2, 0, x, 4000))
#             imgBottom.save(imgBottomName + ".png",  dpi=(300,300), compress_level=2)
